import pandas as pd
import pymongo
import openpyxl
import json

def lastIndex(collection, variable):
    # Use the $group and $max operators in an aggregation pipeline
    pipeline = [
        {
            "$group": {
                "_id": None,
                "max_value": {"$max": variable}
            }
        }
    ]
    result = list(collection.aggregate(pipeline))

    lastIndex = result[0]["max_value"]
    return lastIndex

#vignette index must be compared with lenVignettes



def connect2collection(connection_string, database, collection):
        # Create a connection to MongoDB Atlas
        client = pymongo.MongoClient(connection_string)
        # Select database
        db = client[database]
        # Select prospects within the database
        collection = db[collection]
        return collection


def last_value(collection, variable, n = 1):
    #Obtain last document through ID
    query = collection.find().sort([("_id", pymongo.DESCENDING)]).limit(n)
    if n == 1:
        for i in query:
           result = i.get(variable)
        query = result
    else:
        L = []
        for i in L:
            L.append(i.get(variable))
        query = L
    return query



def export2atlas(list_of_variables, collection):                  
    # Create a dictionary from the data
    data_dict = {col_name: list_of_variables for col_name, list_of_variables in zip(['scraped_var_1', 'scraped_var_i', 'scraped_var_n', 'numeroDEpage', 'indexDANSlaPage' , 'position_dans_la_page', 'totalVignettesInPage', 'jourDuScrapping', 'heureDuScraping', 'capchat_occurrence_since_begening', 'sent', 'VignetteHeight'], list_of_variables)}
    # Create a DataFrame from the dict
    try: 
        df = pd.DataFrame(data_dict)
    except ValueError:
        with open('C:/Users/morel/Documents/Library/CodingTime/GitHub/robinhood/debug.json', 'w') as f:
            json.dump(data_dict, f)
    # Iterate through DataFrame rows and insert each row as a document
    for index, row in df.iterrows():
        document = row.to_dict()  # Convert the row to a dictionary
        collection.insert_one(document)


def import_collection(DB, collection):
    # Connect to the MongoDB collection
    collection = connect2collection(DB, collection)

    # Create a new Excel workbook
    workbook = openpyxl.Workbook()

    # Create a new worksheet and assign it to the worksheet variable
    worksheet = workbook.active

    # Define the headers (column names)
    fields = ['_id', 'scraped_var_1', 'scraped_var_i','jourDuScrapping', 'nomPoste', 'scraped_var_n']

    # Write the headers to the first row of the worksheet
    for col_num, field in enumerate(fields, 1):
        worksheet.cell(row=1, column=col_num, value=field)
    # Retrieve data from MongoDB and write it to the worksheet
    row_num = 2  # Start from the second row
    query = {"sent": 'False'}
    cursor = collection.find(query).limit(500).sort([("$natural", 1)])
    for record in cursor: 
        document_id = record["_id"]
        collection.update_one({"_id": document_id}, {"$set": {"sent": 'True'}})
        for col_num, field in enumerate(fields, 1):
            # Use the field names directly to retrieve values from the record
            worksheet.cell(row=row_num, column=col_num, value=record.get(field, ''))  # Provide a default value if field is missing
        row_num += 1
    return workbook



def import_collection2(DB, collection_name):
    try:
        # Connect to the MongoDB collection
        collection = connect2collection(DB, collection_name)

        # Create a new Excel workbook
        workbook = openpyxl.Workbook()

        # Create a new worksheet and assign it to the worksheet variable
        worksheet = workbook.active

        # Define the headers (column names)
        fields = ['_id', 'scraped_var_1', 'scraped_var_i','jourDuScrapping', 'nomPoste', 'scraped_var_n']

        # Write the headers to the first row of the worksheet
        for col_num, field in enumerate(fields, 1):
            worksheet.cell(row=1, column=col_num, value=field)

        # Retrieve data from MongoDB and write it to the worksheet
        row_num = 2  # Start from the second row
        # query = {"sent": 'False'}
        cursor = collection.find({}).sort([("$natural", 1)])
        for record in cursor: 
            # document_id = record["_id"]
            # collection.update_one({"_id": document_id}, {"$set": {"sent": 'True'}})
            for col_num, field in enumerate(fields, 1):
                # Use the field names directly to retrieve values from the record
                 # Convert ObjectId to string if the field is '_id'
                value = str(record[field]) if field == '_id' else record.get(field, '')
                worksheet.cell(row=row_num, column=col_num, value=value)
            row_num += 1
        return workbook 

    except Exception as e:
        print(f"An error occurred: {e}")



def count_value(collection, column, value):
    pipeline = [
        {
            "$match": {
                column: value
            }
        },
        {
            "$count": "count"
        }
    ]

    result = list(collection.aggregate(pipeline))

    if result:
        count = result[0]["count"]
        return count
    else:
        print("No documents with 'sent' == False found.")
        return 'nada'




